import streamlit as st
import pdfplumber
import pandas as pd
import re
import json
import os
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="7501 Entry Tracker", layout="wide")

st.markdown("""
<style>
  [data-testid="stAppViewContainer"] { background: #0d1b2a; }
  [data-testid="stSidebar"] { background: #0d1b2a; }
  .block-container { padding-top: 1.5rem; }
  h1,h2,h3,h4,h5,h6 { color: #e8f0fe !important; }
  p, label, .stText, div[data-testid="stMarkdownContainer"] p { color: #c9d6e8 !important; }
  [data-testid="metric-container"] { background: #1a2d45; border: 1px solid #2e4a6e; border-radius: 10px; padding: 12px 16px; }
  [data-testid="metric-container"] label { color: #7fa8d4 !important; font-size: 13px !important; }
  [data-testid="metric-container"] [data-testid="stMetricValue"] { color: #e8f0fe !important; font-size: 22px !important; }
  .stTextInput > div > div > input, .stNumberInput > div > div > input { background: #1a2d45 !important; color: #e8f0fe !important; border: 1px solid #2e4a6e !important; border-radius: 6px !important; }
  .stButton > button { background: #1a56db !important; color: white !important; border: none !important; border-radius: 8px !important; padding: 8px 20px !important; font-weight: 600 !important; }
  .stButton > button:hover { background: #1e40af !important; }
  div[data-testid="stExpander"] { background: #1a2d45 !important; border: 1px solid #2e4a6e !important; border-radius: 10px !important; margin-bottom: 8px !important; }
  div[data-testid="stExpander"] summary { color: #e8f0fe !important; }
  .stTabs [data-baseweb="tab-list"] { background: #1a2d45; border-radius: 10px; padding: 4px; }
  .stTabs [data-baseweb="tab"] { color: #7fa8d4 !important; border-radius: 8px !important; }
  .stTabs [aria-selected="true"] { background: #1a56db !important; color: white !important; }
  hr { border-color: #2e4a6e !important; }
  .stFileUploader { background: #1a2d45 !important; border: 2px dashed #2e4a6e !important; border-radius: 10px !important; }
  [data-testid="stCheckbox"] label { color: #c9d6e8 !important; }
  .stMultiSelect span { background: #1a56db !important; color: white !important; }
  .stDateInput input { background: #1a2d45 !important; color: #e8f0fe !important; border: 1px solid #2e4a6e !important; }
  .section-header { background: linear-gradient(90deg,#1a2d45,#0d1b2a); border-left: 4px solid #1a56db; padding: 8px 16px; border-radius: 0 8px 8px 0; margin: 16px 0 12px 0; }
  .section-header p { color: #7fa8d4 !important; font-size: 13px !important; font-weight: 600 !important; letter-spacing: 1px !important; text-transform: uppercase !important; margin: 0 !important; }
</style>
""", unsafe_allow_html=True)

DATA_FILE = "entries.json"

def load_entries():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return []

def save_entries(entries):
    with open(DATA_FILE, "w") as f:
        json.dump(entries, f, indent=2)

def extract_text_from_pdf(pdf_file):
    try:
        with pdfplumber.open(pdf_file) as pdf:
            return "\n".join(p.extract_text() or "" for p in pdf.pages)
    except:
        return ""

def grab(pattern, text, default="", group=1):
    m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return m.group(group).strip() if m else default

COUNTRY_MAP = {
    "CN":"China","MX":"Mexico","BR":"Brazil","DE":"Germany","JP":"Japan",
    "KR":"South Korea","IN":"India","TW":"Taiwan","CA":"Canada","GB":"United Kingdom",
    "FR":"France","IT":"Italy","VN":"Vietnam","TH":"Thailand","MY":"Malaysia","PH":"Philippines",
}

def parse_7501(text):
    entry_number = grab(r'\b(\d{3}-\d{7}-\d)\b', text)
    entry_date   = grab(r'\b\d{3}-\d{7}-\d\b[^\n]+(\d{2}/\d{2}/\d{4})', text)
    carrier_line = re.search(r'CMA CGM[^\n]*\b([A-Z]{2})\b\s+(\d{2}/\d{2}/\d{4})', text)
    if not carrier_line:
        carrier_line = re.search(r'\b([A-Z]{2})\b\s+(\d{2}/\d{2}/\d{4})\s*\n', text)
    country_code = carrier_line.group(1) if carrier_line else ""
    import_date  = carrier_line.group(2) if carrier_line else ""
    country      = COUNTRY_MAP.get(country_code, country_code)
    broker       = grab(r'(KUEHNE\s*\+\s*NAGEL[^\n,\.]*)', text).split('\n')[0].strip()
    if not broker:
        broker = grab(r'Broker/Filer Information[^\n]*\n([^\n]+)', text)
    broker_number = grab(r'\b(BUS\d+)\b', text)
    invoice       = grab(r'Invoice\s+Number\s+(\S+)', text)
    sup_match     = re.search(r'MANUFACTURER\s+SUPPLIER\s*\n([^\n]+)', text)
    if sup_match:
        raw = sup_match.group(1).strip()
        supplier = raw[:len(raw)//2].strip()
    else:
        supplier = grab(r'SUPPLIER\s*\n([^\n]+)', text)
    part_match  = re.search(r'\b(E\d{10}[A-Z0-9]\d)\b', text)
    part_number = part_match.group(1) if part_match else ""
    quantity    = grab(r'\b(\d{3,5})\s+NO\b', text)
    inv_value   = grab(r'Invoice Value USD\s+([\d,]+\.?\d*)', text)
    total_value_str = grab(r'Total Entered Value \(Invoice\)\s+([\d,]+\.?\d*)', text)
    try:
        total_value_float = float(total_value_str.replace(',',''))
    except:
        total_value_float = 0.0
    total_duty_str = grab(r'Total Other Fees\s+([\d,]+\.\d{2})', text)
    try:
        total_duty_float = float(total_duty_str.replace(',',''))
    except:
        total_duty_float = 0.0
    tariff_lines = parse_tariff_lines(text, total_value_float)
    return {
        "entry_number": entry_number, "entry_date": entry_date, "import_date": import_date,
        "broker": broker, "broker_number": broker_number, "supplier": supplier,
        "country": country, "invoice": invoice, "part_number": part_number,
        "quantity": quantity, "invoice_value": inv_value,
        "total_value": total_value_float, "total_duty": total_duty_float,
        "tariff_lines": tariff_lines,
    }

def parse_tariff_lines(text, entered_value):
    lines = []
    seen = set()
    for m in re.finditer(
        r'([0-9]{4}\.[0-9]{2}\.[0-9]{2,4})\s+.*?([0-9]+(?:\.[0-9]+)?)\s*%\s+([0-9,]+\.[0-9]{2})',
        text, re.ASCII
    ):
        hts  = m.group(1)
        rate = float(m.group(2))
        duty = float(m.group(3).replace(',',''))
        if rate < 1.0 or duty == 0.0 or hts in seen:
            continue
        seen.add(hts)
        ev = entered_value
        ev_match = re.search(re.escape(hts)+r'\s+([0-9,]+\.?[0-9]*)\s+\w+\s+([0-9,]+\.?[0-9]*)\s+[0-9]+\s*%', text, re.ASCII)
        if ev_match:
            try:
                c = float(ev_match.group(2).replace(',',''))
                if c > 0: ev = c
            except: pass
        lines.append({"hts":hts,"entered_value":ev,"rate_pct":rate,"duty_amount":duty})
    return lines

def import_from_excel(file):
    try:
        df = pd.read_excel(file, header=1, dtype=str)
        df.columns = [str(c).strip() for c in df.iloc[0]]
        df = df.iloc[1:].reset_index(drop=True)
    except Exception as e:
        return [], [f"Could not read file: {e}"]

    cols = list(df.columns)
    tariff_count = 0
    for i, c in enumerate(cols):
        if c == "Tariff":
            if tariff_count > 0:
                cols[i] = f"Tariff.{tariff_count}"
            tariff_count += 1
    df.columns = cols

    existing = load_entries()
    existing_numbers = {e["entry_number"] for e in existing}
    new_entries = []
    skipped = []

    for _, row in df.iterrows():
        entry_num = str(row.get("Entry #","")).strip()
        if not entry_num or entry_num == "nan":
            continue
        if entry_num in existing_numbers:
            skipped.append(entry_num)
            continue

        def sf(val):
            try:
                v = str(val).replace(',','').replace('%','').strip()
                return float(v) if v and v != "nan" else 0.0
            except: return 0.0

        def ss(val):
            v = str(val).strip()
            return "" if v == "nan" else v

        def fd(val):
            v = str(val).strip()
            if v == "nan" or not v: return ""
            try: return pd.to_datetime(v).strftime("%m/%d/%Y")
            except: return v

        tariff_lines = []
        total_val = sf(row.get("Total", 0))
        for tc, rc, dc in [("Tariff","Rate 1 (%)","Duty 1 (USD)"),("Tariff.1","Rate 2 (%)","Duty 2 (USD)"),("Tariff.2","Rate 3 (%)","Duty 3 (USD)")]:
            hts  = ss(row.get(tc,""))
            rate_raw = sf(row.get(rc, 0))
            rate = rate_raw * 100 if rate_raw < 1 else rate_raw
            duty = sf(row.get(dc, 0))
            if hts and rate >= 1.0 and duty > 0:
                tariff_lines.append({"hts":hts,"entered_value":total_val,"rate_pct":rate,"duty_amount":duty})

        total_duty = sf(row.get("Total Duty (USD)", 0))
        tdp        = sf(row.get("Total Duty %", 0))
        eff_rate   = round(tdp * 100, 2) if tdp < 1 else round(tdp, 2)

        new_entries.append({
            "date_logged":   datetime.now().strftime("%Y-%m-%d %H:%M"),
            "entry_number":  entry_num,
            "entry_date":    fd(row.get("Entry Date","")),
            "import_date":   fd(row.get("Import Date","")),
            "broker":        ss(row.get("Broker","")),
            "broker_number": ss(row.get("Broker number","")),
            "supplier":      ss(row.get("Supplier Name","")),
            "country":       ss(row.get("Country of Origin","")),
            "invoice":       ss(row.get("Invoice #","")),
            "part_number":   ss(row.get("Part Number","")),
            "quantity":      ss(row.get("Quantity (pcs)","")),
            "invoice_value": ss(row.get("Price","")),
            "total_value":   total_val,
            "total_duty":    total_duty,
            "eff_rate":      eff_rate,
            "tariff_lines":  tariff_lines,
            "filename":      "imported from Excel",
        })
        existing_numbers.add(entry_num)

    return new_entries, skipped

def build_excel(entries):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "7501 Log"
    hfill = PatternFill("solid", start_color="1a2d45")
    hfont = Font(name="Arial", bold=True, color="7fa8d4", size=10)
    hal   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dfont = Font(name="Arial", size=10)
    headers = ["Entry #","Entry Date","Import Date","Broker","Broker #","Supplier","Country",
               "Invoice #","Part #","Qty","Invoice Value","Total Value","HTS Code",
               "Entered Value","Rate (%)","Duty (USD)","Total Duty","Eff Rate %","Date Logged"]
    widths  = [16,12,12,20,14,22,14,18,18,10,14,14,14,14,10,14,14,12,16]
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(1,c,h); cell.font=hfont; cell.fill=hfill; cell.alignment=hal
        ws.column_dimensions[cell.column_letter].width=w
    ws.row_dimensions[1].height=30
    row=2
    for e in entries:
        tlines = e.get("tariff_lines",[]) or [{"hts":"","entered_value":"","rate_pct":"","duty_amount":""}]
        for t in tlines:
            vals=[e.get("entry_number"),e.get("entry_date"),e.get("import_date"),
                  e.get("broker"),e.get("broker_number"),e.get("supplier"),e.get("country"),
                  e.get("invoice"),e.get("part_number"),e.get("quantity"),e.get("invoice_value"),
                  e.get("total_value"),t.get("hts"),t.get("entered_value"),t.get("rate_pct"),
                  t.get("duty_amount"),e.get("total_duty"),e.get("eff_rate"),e.get("date_logged")]
            for c,v in enumerate(vals,1):
                cell=ws.cell(row,c,v); cell.font=dfont
                cell.alignment=Alignment(horizontal="center",vertical="center")
            row+=1
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

if "tariff_lines" not in st.session_state:
    st.session_state.tariff_lines = [{}]

st.markdown("""
<div style="background:linear-gradient(90deg,#1a2d45,#0d1b2a);padding:20px 24px;border-radius:12px;margin-bottom:24px;border:1px solid #2e4a6e;">
  <h1 style="margin:0;color:#e8f0fe;font-size:26px;">📦 CBP Form 7501 — Entry Tracker</h1>
  <p style="margin:4px 0 0;color:#7fa8d4;font-size:13px;">BorgWarner Dixon LLC  •  Customs Entry Management</p>
</div>
""", unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs(["➕  New Entry", "📂  Import from Excel", "📋  View All Entries"])

# ── TAB 1: NEW ENTRY ──────────────────────────────────────────────────────────
with tab1:
    uploaded = st.file_uploader("Upload 7501 PDF (optional — fields will auto-fill)", type="pdf")
    d = {}
    if uploaded:
        text = extract_text_from_pdf(uploaded)
        d = parse_7501(text)
        st.session_state.tariff_lines = d.get("tariff_lines") or [{}]
        st.success("✅ PDF parsed — review and correct fields below before saving.")

    st.markdown('<div class="section-header"><p>Entry / Filing Info</p></div>', unsafe_allow_html=True)
    c1,c2,c3,c4 = st.columns(4)
    with c1:
        entry_number  = st.text_input("Entry #",           value=d.get("entry_number",""))
        broker        = st.text_input("Broker",            value=d.get("broker",""))
    with c2:
        entry_date    = st.text_input("Entry Date",        value=d.get("entry_date",""))
        broker_number = st.text_input("Broker #",          value=d.get("broker_number",""))
    with c3:
        import_date   = st.text_input("Import Date",       value=d.get("import_date",""))
        supplier      = st.text_input("Supplier",          value=d.get("supplier",""))
    with c4:
        country       = st.text_input("Country of Origin", value=d.get("country",""))
        invoice       = st.text_input("Invoice #",         value=d.get("invoice",""))

    st.markdown('<div class="section-header"><p>Product / Part</p></div>', unsafe_allow_html=True)
    p1,p2,p3 = st.columns(3)
    with p1: part_number   = st.text_input("Part Number",         value=d.get("part_number",""))
    with p2: quantity      = st.text_input("Quantity",            value=d.get("quantity",""))
    with p3: invoice_value = st.text_input("Invoice Value (USD)", value=d.get("invoice_value",""))

    st.markdown('<div class="section-header"><p>Tariff Lines — charges ≥ 1% only</p></div>', unsafe_allow_html=True)
    th1,th2,th3,th4 = st.columns([2.5,2,1.5,2])
    th1.markdown("**HTS Code**"); th2.markdown("**Entered Value ($)**")
    th3.markdown("**Rate (%)**"); th4.markdown("**Duty (USD)**")

    updated_tariffs = []
    for i, t in enumerate(st.session_state.tariff_lines):
        tc1,tc2,tc3,tc4 = st.columns([2.5,2,1.5,2])
        hts = tc1.text_input("", value=str(t.get("hts","") or ""),           key=f"hts_{i}", label_visibility="collapsed")
        ev  = tc2.number_input("",value=float(t.get("entered_value",0) or 0),key=f"ev_{i}",  label_visibility="collapsed", min_value=0.0, format="%.2f")
        rt  = tc3.number_input("",value=float(t.get("rate_pct",0) or 0),     key=f"rt_{i}",  label_visibility="collapsed", min_value=0.0, max_value=100.0, format="%.2f")
        dy  = tc4.number_input("",value=float(t.get("duty_amount",0) or 0),  key=f"dy_{i}",  label_visibility="collapsed", min_value=0.0, format="%.2f")
        updated_tariffs.append({"hts":hts,"entered_value":ev,"rate_pct":rt,"duty_amount":dy})
    st.session_state.tariff_lines = updated_tariffs

    ba,bb,_ = st.columns([1,1,5])
    if ba.button("➕ Add Line"):
        st.session_state.tariff_lines.append({})
        st.rerun()
    if bb.button("🗑 Remove Last") and len(st.session_state.tariff_lines) > 1:
        st.session_state.tariff_lines.pop()
        st.rerun()

    st.markdown('<div class="section-header"><p>Totals</p></div>', unsafe_allow_html=True)
    tv_input = st.number_input("Total Entered Value (USD)", value=float(d.get("total_value",0) or 0), min_value=0.0, format="%.2f")
    td_input = st.number_input("Total Duty (USD)",          value=float(d.get("total_duty",0)  or 0), min_value=0.0, format="%.2f")
    eff_rate = round((td_input/tv_input*100) if tv_input else 0, 2)
    s1,s2,s3 = st.columns(3)
    s1.metric("Total Entered Value", f"${tv_input:,.2f}")
    s2.metric("Total Duty",          f"${td_input:,.2f}")
    s3.metric("Effective Duty Rate", f"{eff_rate:.2f}%")

    st.markdown("---")
    if st.button("✅ Save Entry", type="primary"):
        if not entry_number:
            st.error("Entry number is required.")
        else:
            entries = load_entries()
            if entry_number in {e["entry_number"] for e in entries}:
                st.warning(f"Entry {entry_number} already exists.")
            else:
                entries.append({
                    "date_logged": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "entry_number": entry_number, "entry_date": entry_date,
                    "import_date": import_date, "broker": broker,
                    "broker_number": broker_number, "supplier": supplier,
                    "country": country, "invoice": invoice,
                    "part_number": part_number, "quantity": quantity,
                    "invoice_value": invoice_value, "total_value": tv_input,
                    "total_duty": td_input, "eff_rate": eff_rate,
                    "tariff_lines": [t for t in st.session_state.tariff_lines if t.get("hts")],
                    "filename": uploaded.name if uploaded else "",
                })
                save_entries(entries)
                st.session_state.tariff_lines = [{}]
                st.success(f"✅ Entry {entry_number} saved!")
                st.rerun()

# ── TAB 2: IMPORT FROM EXCEL ──────────────────────────────────────────────────
with tab2:
    st.markdown('<div class="section-header"><p>Bulk Import from Excel</p></div>', unsafe_allow_html=True)
    st.markdown("Upload your existing CBP 7501 Excel tracker. Duplicate entry numbers are skipped automatically.")

    xl_file = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx","xls"], key="xl_upload")

    if xl_file:
        with st.spinner("Reading Excel file..."):
            new_entries, skipped = import_from_excel(xl_file)

        st.markdown("---")
        ci1,ci2 = st.columns(2)
        ci1.metric("New entries found",  len(new_entries))
        ci2.metric("Duplicates skipped", len(skipped))

        if new_entries:
            st.markdown("**Preview — first 5 entries to be imported:**")
            preview = []
            for e in new_entries[:5]:
                preview.append({
                    "Entry #":     e["entry_number"],
                    "Entry Date":  e["entry_date"],
                    "Broker":      e["broker"],
                    "Supplier":    e["supplier"],
                    "Country":     e["country"],
                    "Total Value": f"${e['total_value']:,.2f}",
                    "Total Duty":  f"${e['total_duty']:,.2f}",
                    "Tariff Lines":len(e["tariff_lines"]),
                })
            st.dataframe(preview, use_container_width=True)

            if skipped:
                st.warning(f"Already exist (will skip): {', '.join(skipped[:10])}{'...' if len(skipped)>10 else ''}")

            if st.button(f"📥 Import {len(new_entries)} entries", type="primary"):
                existing = load_entries()
                existing.extend(new_entries)
                save_entries(existing)
                st.success(f"✅ Successfully imported {len(new_entries)} entries!")
                st.rerun()
        elif not skipped:
            st.info("No valid entries found in the file.")
        else:
            st.warning("All entries in this file already exist in the tracker.")

# ── TAB 3: VIEW ALL ───────────────────────────────────────────────────────────
with tab3:
    entries = load_entries()

    if not entries:
        st.info("No entries logged yet. Add one in the New Entry tab or import from Excel.")
    else:
        total_entries = len(entries)
        total_value   = sum(e.get("total_value",0) for e in entries)
        total_duty    = sum(e.get("total_duty",0)  for e in entries)
        avg_rate      = round((total_duty/total_value*100) if total_value else 0, 2)

        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Total Entries",       total_entries)
        m2.metric("Total Entered Value", f"${total_value:,.2f}")
        m3.metric("Total Duties Paid",   f"${total_duty:,.2f}")
        m4.metric("Avg Effective Rate",  f"{avg_rate:.2f}%")

        st.markdown("---")
        with st.expander("🔽  Filters & Search", expanded=True):
            fc1,fc2,fc3 = st.columns(3)
            with fc1: search = st.text_input("🔍 Search", placeholder="Entry #, supplier, part #, HTS, country, broker...")
            with fc2:
                all_brokers = sorted(set(e.get("broker","") for e in entries if e.get("broker")))
                broker_filter = st.multiselect("Broker", all_brokers)
            with fc3:
                all_countries = sorted(set(e.get("country","") for e in entries if e.get("country")))
                country_filter = st.multiselect("Country of Origin", all_countries)

            fc4,fc5,fc6 = st.columns(3)
            with fc4:
                all_suppliers = sorted(set(e.get("supplier","") for e in entries if e.get("supplier")))
                supplier_filter = st.multiselect("Supplier", all_suppliers)
            with fc5: date_from = st.date_input("Entry Date From", value=None)
            with fc6: date_to   = st.date_input("Entry Date To",   value=None)

            fc7,fc8 = st.columns(2)
            with fc7: min_duty = st.number_input("Min Total Duty ($)", value=0.0, min_value=0.0, format="%.2f")
            with fc8: max_duty = st.number_input("Max Total Duty ($)", value=0.0, min_value=0.0, format="%.2f", help="Leave at 0 for no max")

        filtered = entries
        if search:
            s = search.lower()
            filtered = [e for e in filtered if
                s in e.get("entry_number","").lower() or s in e.get("supplier","").lower() or
                s in e.get("country","").lower() or s in e.get("part_number","").lower() or
                s in e.get("broker","").lower() or s in e.get("invoice","").lower() or
                any(s in t.get("hts","").lower() for t in e.get("tariff_lines",[]))]
        if broker_filter:   filtered = [e for e in filtered if e.get("broker") in broker_filter]
        if country_filter:  filtered = [e for e in filtered if e.get("country") in country_filter]
        if supplier_filter: filtered = [e for e in filtered if e.get("supplier") in supplier_filter]
        if date_from: filtered = [e for e in filtered if e.get("entry_date","") >= date_from.strftime("%m/%d/%Y")]
        if date_to:   filtered = [e for e in filtered if e.get("entry_date","") <= date_to.strftime("%m/%d/%Y")]
        if min_duty > 0: filtered = [e for e in filtered if e.get("total_duty",0) >= min_duty]
        if max_duty > 0: filtered = [e for e in filtered if e.get("total_duty",0) <= max_duty]

        st.markdown(f"Showing **{len(filtered)}** of **{total_entries}** entries")

        to_delete = []
        for i, e in enumerate(reversed(filtered)):
            eff = e.get("eff_rate", 0)
            label = (f"📄  {e['entry_number']}  │  {e.get('supplier','')}  │  "
                     f"{e.get('country','')}  │  {e.get('entry_date','')}  │  "
                     f"${e.get('total_value',0):,.2f}  │  Duty: ${e.get('total_duty',0):,.2f} ({eff:.1f}%)")

            col_chk, col_exp = st.columns([0.04, 0.96])
            if col_chk.checkbox("", key=f"del_{i}_{e['entry_number']}"):
                to_delete.append(e["entry_number"])

            with col_exp.expander(label):
                h1,h2,h3,h4 = st.columns(4)
                h1.markdown(f"**Entry #:** {e.get('entry_number')}")
                h1.markdown(f"**Entry Date:** {e.get('entry_date')}")
                h1.markdown(f"**Import Date:** {e.get('import_date')}")
                h2.markdown(f"**Broker:** {e.get('broker')}")
                h2.markdown(f"**Broker #:** {e.get('broker_number')}")
                h2.markdown(f"**Invoice #:** {e.get('invoice')}")
                h3.markdown(f"**Supplier:** {e.get('supplier')}")
                h3.markdown(f"**Country:** {e.get('country')}")
                h3.markdown(f"**Part #:** {e.get('part_number')}")
                h4.markdown(f"**Quantity:** {e.get('quantity')}")
                h4.markdown(f"**Invoice Value:** {e.get('invoice_value','')}")
                h4.markdown(f"**Logged:** {e.get('date_logged')}")

                tariff_lines = e.get("tariff_lines",[])
                if tariff_lines:
                    st.markdown("**Tariff Lines:**")
                    tl1,tl2,tl3,tl4 = st.columns([2.5,2,1.5,2])
                    tl1.markdown("**HTS Code**"); tl2.markdown("**Entered Value**")
                    tl3.markdown("**Rate**");     tl4.markdown("**Duty**")
                    for t in tariff_lines:
                        tl1,tl2,tl3,tl4 = st.columns([2.5,2,1.5,2])
                        tl1.write(t.get("hts",""))
                        tl2.write(f"${float(t.get('entered_value',0) or 0):,.2f}")
                        tl3.write(f"{float(t.get('rate_pct',0) or 0):.1f}%")
                        tl4.write(f"${float(t.get('duty_amount',0) or 0):,.2f}")

                st.markdown("---")
                sv1,sv2,sv3 = st.columns(3)
                sv1.metric("Total Entered Value", f"${e.get('total_value',0):,.2f}")
                sv2.metric("Total Duty",          f"${e.get('total_duty',0):,.2f}")
                sv3.metric("Effective Rate",       f"{eff:.2f}%")

        if to_delete:
            st.warning(f"{len(to_delete)} entry/entries selected for deletion.")
            if st.button("🗑 Delete Selected", type="primary"):
                updated = [e for e in entries if e["entry_number"] not in to_delete]
                save_entries(updated)
                st.success(f"Deleted {len(to_delete)} entries.")
                st.rerun()

        st.markdown("---")
        if st.button("📥 Export All to Excel"):
            buf = build_excel(entries)
            st.download_button(
                label="⬇️ Download Excel",
                data=buf,
                file_name=f"7501_log_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
